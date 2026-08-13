"""
app.py

Production-facing Flask API for the spreadsheet-to-dashboard pipeline.

Key production concerns handled here:
  - File size limit (MAX_CONTENT_LENGTH) -- prevents huge uploads
  - Row count cap -- prevents one file from generating a huge/expensive
    Claude API call or slow processing
  - Excel (.xlsx) support in addition to .csv -- real small business
    spreadsheets are usually Excel, not CSV
  - Rate limiting per IP -- prevents one user (or abuse/bots) from
    silently running up your API bill
  - CORS restricted to your actual frontend domain in production
  - The raw uploaded file is NEVER written to disk -- it's parsed
    in-memory and discarded. Worth stating this plainly to users as a
    privacy note on the landing page.

Run locally:
  export ANTHROPIC_API_KEY="your-key-here"
  pip install -r requirements.txt
  python app.py

Run in production: see DEPLOYMENT.md
"""

import os
import io
import json
import time
import uuid
import pandas as pd
from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

from type_inference import infer_all_columns, apply_column_overrides, _clean_numeric_series
from chart_engine import build_charts
from qa_engine import parse_question
from data_quality import build_data_quality_report
from semantic_labeler import understand_columns
from summary_writer import write_summaries
from events_logger import log_event, get_stats

app = Flask(__name__)

# --- Config -----------------------------------------------------------------
MAX_FILE_SIZE_MB = 5
MAX_ROWS = 20_000
MAX_COLUMNS = 500
FRONTEND_ORIGIN = os.environ.get("FRONTEND_ORIGIN", "*")  # set this in production!
UPLOAD_RATE_LIMIT = os.environ.get("UPLOAD_RATE_LIMIT", "10 per hour")

app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_SIZE_MB * 1024 * 1024

CORS(app, origins=[FRONTEND_ORIGIN] if FRONTEND_ORIGIN != "*" else "*")

limiter = Limiter(get_remote_address, app=app, default_limits=[])


def _detect_header_row(raw_df: pd.DataFrame, scan_rows: int = 15) -> int:
    """
    Real-world spreadsheets often have title rows, blank spacer rows, or
    a company logo row sitting above the actual column headers -- pandas'
    default assumption (row 1 = headers) fails on these, producing
    'Unnamed: 0', 'Unnamed: 1', etc for every column.

    This scans the first `scan_rows` rows and picks the one with the most
    non-null values as the real header row -- a title row typically has
    only 1-2 non-null cells (just the title text), while a real header row
    has one non-null cell per column.
    """
    n_scan = min(scan_rows, len(raw_df))
    non_null_counts = raw_df.iloc[:n_scan].notna().sum(axis=1)
    return int(non_null_counts.idxmax())


class PipelineError(ValueError):
    """
    A ValueError that also carries a specific, named error code -- so the
    response and the logs say exactly what went wrong (NO_TABULAR_DATA,
    UNSUPPORTED_FILE_TYPE, ...) instead of every failure looking the same
    behind one generic message. Stays a ValueError so it's still caught
    by any existing `except ValueError` handling.
    """
    def __init__(self, message, code):
        super().__init__(message)
        self.code = code


def _list_excel_sheets(content: bytes) -> list:
    """
    Cheaply lists sheet names and approximate sizes WITHOUT loading full
    data into pandas -- uses openpyxl's read_only mode so this stays fast
    even on a large file. Used to detect multi-sheet workbooks before
    committing to analyzing just one sheet.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({
            "name": name,
            "n_rows": max(0, (ws.max_row or 1) - 1),
            "n_cols": ws.max_column or 0,
        })
    wb.close()
    return sheets


def _read_spreadsheet_from_content(content: bytes, filename: str, sheet_name: str = None) -> pd.DataFrame:
    filename = (filename or "").lower()

    if filename.endswith(".csv"):
        try:
            raw = pd.read_csv(io.BytesIO(content), header=None)
        except pd.errors.ParserError:
            # A real CSV has a consistent number of fields per line. This
            # specific error means the file doesn't -- almost always
            # because it's a formatted report or document, not a table
            # (e.g. "Revenue: $40,000" has a comma inside the number,
            # while blank/title lines around it have none at all).
            raise PipelineError(
                "This looks like a formatted report or document rather than structured "
                "tabular data. Try uploading the underlying data with one consistent row "
                "of column headers, or export it as a plain spreadsheet.",
                code="NO_TABULAR_DATA",
            )
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            raw = pd.read_excel(io.BytesIO(content), header=None, sheet_name=sheet_name if sheet_name else 0)
        except Exception:
            # Common real-world mistake: a CSV file saved or renamed with an
            # Excel extension. Try reading it as CSV before giving up --
            # this is a genuine recovery, not just a nicer error message.
            try:
                raw = pd.read_csv(io.BytesIO(content), header=None)
            except Exception:
                raise PipelineError(
                    "This file has an Excel extension but doesn't appear to be a "
                    "valid .xls/.xlsx file, and it isn't readable as a CSV either. "
                    "Try re-exporting it from its original source.",
                    code="CORRUPTED_FILE",
                )
    else:
        raise PipelineError("Unsupported file type. Please upload a .csv or .xlsx file.", code="UNSUPPORTED_FILE_TYPE")

    # A single-column result (no delimiters ever found) is the other
    # common shape of "not really a dataset" -- e.g. a document where
    # every line happens to have the same (zero) comma count, so the
    # parser above didn't even error.
    if len(raw.columns) <= 1:
        raise PipelineError(
            "This file doesn't appear to contain multiple columns of data. "
            "Try uploading a spreadsheet with distinct columns like Date, Amount, Category, etc.",
            code="NO_TABULAR_DATA",
        )

    header_row = _detect_header_row(raw)
    df = raw.iloc[header_row + 1:].copy()
    df.columns = raw.iloc[header_row]
    df = df.reset_index(drop=True)

    # Drop fully-blank columns (common when a leading/trailing spacer
    # column has no header and no data at all -- pure visual padding
    # in the original spreadsheet, not a real column).
    df = df.dropna(axis=1, how="all")

    # Any remaining column with no real header text still needs a name
    # so nothing downstream breaks on a NaN/blank column label.
    df.columns = [
        str(c) if pd.notna(c) and str(c).strip() else f"Column {i+1}"
        for i, c in enumerate(df.columns)
    ]

    return df


SAMPLE_ROWS_CAP = 15000  # matches MAX_ROWS below in practice -- most real files
# get their FULL data for exploration, not a sample. Only files with more
# rows than this (rare for small-business use cases) get capped, purely
# to keep the JSON payload and browser-side computation reasonably fast.
CHARTABLE_TYPES = {"numeric_measure", "date", "categorical", "boolean"}
MAX_FREE_TEXT_AVG_LEN_FOR_SAMPLE = 60


def _lookupable_columns(df: pd.DataFrame, column_types: dict) -> dict:
    result = {c: t["type"] for c, t in column_types.items() if t["type"] in CHARTABLE_TYPES}
    for col, info in column_types.items():
        if info["type"] == "identifier":
            result[col] = "identifier"
        elif info["type"] == "free_text":
            sample = df[col].dropna().astype(str).head(200)
            if len(sample) == 0 or sample.str.len().mean() <= MAX_FREE_TEXT_AVG_LEN_FOR_SAMPLE:
                result[col] = "free_text"
    return result


def _build_row_sample(df: pd.DataFrame, column_types: dict) -> tuple[list, dict]:
    """
    Returns (sample_rows, chartable_columns) for client-side interactive
    charting (histograms, scatter plots, pie charts, column toggling).
    Only includes columns with a chartable type -- free text, identifiers,
    and empty/constant columns are excluded, both because they're not
    useful to chart and to keep the JSON payload reasonably small.

    Processes each column ONCE, vectorized across all rows -- not once per
    individual cell. An earlier version called pd.to_datetime() and
    _clean_numeric_series() on a single value at a time inside a Python
    loop over every row. That's fine for a few hundred rows in local
    testing, but a real production file with thousands of rows times
    several columns meant tens of thousands of expensive per-value pandas
    calls -- which caused an actual timeout and out-of-memory worker kill
    on Render. This version does the same cleaning, just column-at-a-time.
    """
    chartable_columns = _lookupable_columns(df, column_types)
    if not chartable_columns:
        return [], {}

    sample_df = df[list(chartable_columns.keys())]
    if len(sample_df) > SAMPLE_ROWS_CAP:
        sample_df = sample_df.sample(SAMPLE_ROWS_CAP, random_state=42)

    cleaned = pd.DataFrame(index=sample_df.index)
    for col, col_type in chartable_columns.items():
        if col_type == "date":
            parsed = pd.to_datetime(sample_df[col], errors="coerce", format="mixed")
            cleaned[col] = parsed.dt.strftime("%Y-%m-%d")
        elif col_type == "numeric_measure":
            cleaned[col] = _clean_numeric_series(sample_df[col], col)
        else:
            cleaned[col] = sample_df[col].astype(str)
            cleaned.loc[sample_df[col].isna(), col] = None

    # NaN/NaT survive the column-by-column cleaning above in numeric and
    # date columns. json.dumps does NOT error on a raw float NaN, but it
    # emits the literal text "NaN" in the output -- which is invalid JSON
    # (not part of the spec) and would break JSON.parse() in the browser.
    # Converting to an object-dtype frame and replacing NaN with a real
    # None produces a proper "null" instead. Verified this distinction
    # directly before trusting it, given what shipping this wrong would
    # have caused.
    cleaned = cleaned.astype(object).where(cleaned.notna(), None)

    records = cleaned.to_dict(orient="records")
    # numpy scalar types (np.float64 etc) can linger after the above;
    # convert numeric columns to native Python floats explicitly so the
    # JSON encoder never has to guess.
    numeric_cols = [c for c, t in chartable_columns.items() if t == "numeric_measure"]
    for r in records:
        for col in numeric_cols:
            if r[col] is not None:
                r[col] = float(r[col])

    return records, chartable_columns


@app.route("/api/analyze", methods=["POST"])
@limiter.limit(UPLOAD_RATE_LIMIT)
def analyze():
    upload_id = str(uuid.uuid4())
    start_time = time.time()

    if "file" not in request.files:
        log_event("upload_failed", upload_id=upload_id, reason="NO_FILE_PROVIDED")
        return jsonify({"error": "No file provided", "error_code": "NO_FILE_PROVIDED"}), 400

    file = request.files["file"]
    filename = file.filename or "unknown"
    requested_sheet = request.form.get("sheet_name")
    content = file.read()

    # If this is a multi-sheet Excel file and the frontend hasn't told us
    # which sheet to analyze yet, stop here and ask -- don't silently pick
    # sheet 0 and hope it's the right one. A real business workbook often
    # has a "Sales" sheet, a "Customers" sheet, and an "Instructions" sheet
    # that should never be analyzed as if it were data.
    if filename.lower().endswith((".xlsx", ".xls")) and not requested_sheet:
        try:
            sheets = _list_excel_sheets(content)
        except Exception:
            sheets = []
        if len(sheets) > 1:
            log_event("multiple_sheets_detected", upload_id=upload_id, filename=filename, n_sheets=len(sheets))
            return jsonify({"multiple_sheets": True, "sheets": sheets, "filename": filename})

    try:
        df = _read_spreadsheet_from_content(content, filename, sheet_name=requested_sheet)
    except PipelineError as e:
        log_event("upload_failed", upload_id=upload_id, reason=e.code, filename=filename)
        return jsonify({"error": str(e), "error_code": e.code}), 400
    except Exception as e:
        log_event("upload_failed", upload_id=upload_id, reason="PARSE_ERROR", detail=str(e))
        return jsonify({
            "error": f"Could not parse file. Make sure it's a valid spreadsheet. ({e})",
            "error_code": "PARSE_ERROR",
        }), 400

    if df.empty or len(df.columns) == 0:
        log_event("upload_failed", upload_id=upload_id, reason="EMPTY_DATASET")
        return jsonify({"error": "File appears to be empty", "error_code": "EMPTY_DATASET"}), 400

    if len(df.columns) > MAX_COLUMNS:
        log_event("upload_failed", upload_id=upload_id, reason="TOO_MANY_COLUMNS", n_cols=len(df.columns))
        return jsonify({
            "error": f"This file has {len(df.columns):,} columns, more than the {MAX_COLUMNS:,} limit. "
                     f"Try splitting it into smaller files, or removing columns you don't need analyzed.",
            "error_code": "TOO_MANY_COLUMNS",
        }), 400

    truncated = False
    if len(df) > MAX_ROWS:
        df = df.head(MAX_ROWS)
        truncated = True

    log_event("upload_succeeded", upload_id=upload_id, filename=filename,
               n_rows=len(df), n_cols=len(df.columns))

    # EVERYTHING below this point processes real, possibly messy user data
    # through several independent stages (type inference, chart building,
    # AI calls, sampling). Any ONE of those could theoretically raise on a
    # dataset shape nobody anticipated. This blanket try/except is the
    # final backstop: no matter what goes wrong inside, the person gets a
    # clean error message, never a raw 500 with a Python traceback.
    try:
        # 1. Determine each column's statistical type -- deterministic, no AI,
        #    works identically regardless of what the spreadsheet is about.
        column_types = infer_all_columns(df)

        # 1b. Data quality report -- missing values, duplicate rows, and
        #     IQR-based outliers. Entirely deterministic statistics, no AI,
        #     so this never depends on an API key either.
        data_quality_report = build_data_quality_report(df, column_types)

        # 2. Let AI review EVERY numeric column's aggregation, suggest a
        #    clean label for every column, and write a short overview of what
        #    the dataset appears to be about -- one batched call. Graceful
        #    degradation: if this fails entirely (no API key, API outage), we
        #    fall back to keyword-based agg_kind, title-cased names, and no
        #    overview text below -- the dashboard still works, just less richly.
        numeric_defaults = {
            col: info.get("agg_kind", "sum")
            for col, info in column_types.items() if info.get("type") == "numeric_measure"
        }
        samples = {col: df[col].dropna().head(5).astype(str).tolist() for col in df.columns}
        overview_sample_rows = df.head(5).astype(str).to_dict(orient="records")

        try:
            understood = understand_columns(list(df.columns), samples, numeric_defaults, overview_sample_rows, n_rows=len(df))
        except Exception as e:
            log_event("column_understanding_failed", upload_id=upload_id, detail=str(e))
            understood = {}
        if not understood:
            log_event("column_understanding_unavailable", upload_id=upload_id)

        understood_columns = understood.get("columns", {})
        dataset_overview = understood.get("overview", "")
        # Defensive check: never trust AI's stated row count even with explicit
        # instructions not to guess it. If the real count doesn't appear
        # anywhere in the generated text, don't risk showing a wrong number.
        if dataset_overview and str(len(df)) not in dataset_overview.replace(",", ""):
            dataset_overview = f"This dataset has {len(df):,} rows. " + dataset_overview

        labels = {}
        for col in df.columns:
            info = understood_columns.get(col, {})
            labels[col] = info.get("label") or str(col).replace("_", " ").replace("-", " ").title()
            if column_types.get(col, {}).get("type") == "numeric_measure":
                ai_kind = info.get("agg_kind")
                if ai_kind in ("sum", "average"):
                    column_types[col]["agg_kind"] = ai_kind

        # 2b. Apply any user corrections LAST, after both the statistical
        #     defaults and the AI review above -- a human correction must
        #     always win, with no exceptions, or this feature is pointless.
        #     Recomputed with the same statistical helpers as auto-detection,
        #     so a corrected column is never treated as second-class.
        column_overrides_raw = request.form.get("column_overrides")
        if column_overrides_raw:
            try:
                column_overrides = json.loads(column_overrides_raw)
                column_types = apply_column_overrides(df, column_types, column_overrides)
                log_event("column_overrides_applied", upload_id=upload_id, overridden_columns=list(column_overrides.keys()))
            except Exception as e:
                log_event("column_overrides_failed", upload_id=upload_id, detail=str(e))
                # Don't fail the whole request over a malformed override --
                # just proceed with the auto-detected types instead.

        # 3. Build charts from the statistical types -- this step has NO
        #    dependency on AI at all, so it can never fail because of an
        #    API outage or a bad AI guess.
        dashboard = build_charts(df, column_types, labels)

        # 3a. If near-duplicate entities were found (same real-world record
        #     repeated with slightly different data -- e.g. the same app
        #     scraped on different days), warn on every SUM-type chart,
        #     since summing across duplicated entities silently inflates
        #     the total. Appends to any existing warning rather than
        #     overwriting it, since a chart could have both an outlier
        #     warning and a duplicate-entity warning at once.
        if data_quality_report.get("duplicate_entities"):
            dup_col = next(iter(data_quality_report["duplicate_entities"]))
            dup_info = data_quality_report["duplicate_entities"][dup_col]
            dup_warning = (
                f"{dup_info['n_affected_rows']} rows appear to be the same real-world "
                f"record repeated with slightly different data (grouped by '{labels.get(dup_col, dup_col)}'). "
                f"This total may be inflated as a result. See Data Quality for details."
            )
            for c in dashboard["charts"]:
                measure_col = c.get("measure_column")
                is_sum_chart = (
                    (c["type"] == "summary_card" and c.get("agg_kind") == "sum")
                    or (c["type"] in ("bar", "line") and measure_col and column_types.get(measure_col, {}).get("agg_kind") == "sum")
                )
                if is_sum_chart:
                    if c.get("warning"):
                        c["warning"] = c["warning"] + " " + dup_warning
                    else:
                        c["warning"] = dup_warning

        # 3b. Ask AI to phrase each chart's already-computed facts as a plain
        #     sentence -- purely cosmetic, same principle as the labeler. If
        #     this fails, the frontend just shows the stat_summary text as-is.
        chartable_summaries = [
            {"title": c["title"], "facts": c["summary"]["stat_summary"]}
            for c in dashboard["charts"]
            if c.get("summary", {}).get("stat_summary")
        ]
        try:
            sentences = write_summaries(chartable_summaries)
        except Exception as e:
            log_event("summary_writing_failed", upload_id=upload_id, detail=str(e))
            sentences = {}
        for c in dashboard["charts"]:
            if c.get("summary", {}).get("stat_summary"):
                c["summary"]["ai_sentence"] = sentences.get(c["title"])

        # 4. Sample raw rows for client-side interactive exploration (scatter
        #    plots, histograms, pie charts, user-chosen column combinations) --
        #    separate from the auto-generated charts above.
        sample_rows, chartable_columns = _build_row_sample(df, column_types)

        elapsed_ms = round((time.time() - start_time) * 1000, 1)
        log_event(
            "dashboard_rendered",
            upload_id=upload_id,
            elapsed_ms=elapsed_ms,
            n_charts=len(dashboard["charts"]),
            chart_types=[c["type"] for c in dashboard["charts"]],
        )

        return jsonify({
            "upload_id": upload_id,
            "n_rows": len(df),
            "n_cols": len(df.columns),
            "truncated": truncated,
            "max_rows": MAX_ROWS,
            "column_types": {col: info["type"] for col, info in column_types.items()},
            "column_agg_kinds": {
                col: info.get("agg_kind", "sum")
                for col, info in column_types.items() if info["type"] == "numeric_measure"
            },
            "column_confidence": {
                col: {
                    "type_confidence": info.get("type_confidence", 1.0),
                    "agg_kind_confidence": info.get("agg_kind_confidence"),
                }
                for col, info in column_types.items()
            },
            "user_corrected_columns": [col for col, info in column_types.items() if info.get("user_corrected")],
            "correction_notes": {
                col: info["correction_note"] for col, info in column_types.items() if info.get("correction_note")
            },
            "labels": labels,
            "dataset_overview": dataset_overview,
            "data_quality": data_quality_report,
            "overview": dashboard["overview"],
            "charts": dashboard["charts"],
            "sample_rows": sample_rows,
            "chartable_columns": chartable_columns,
            "elapsed_ms": elapsed_ms,
        })
    except Exception as e:
        log_event("analysis_failed_unexpectedly", upload_id=upload_id, filename=filename, detail=str(e))
        return jsonify({
            "error": "Something unexpected went wrong analyzing this file. "
                      "This has been logged. Try a different file, or try this one again.",
            "error_code": "DASHBOARD_GENERATION_FAILED",
        }), 500


@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": f"File too large. Max size is {MAX_FILE_SIZE_MB}MB.", "error_code": "FILE_TOO_LARGE"}), 413


@app.errorhandler(429)
def rate_limited(e):
    return jsonify({"error": "Too many uploads. Please try again later.", "error_code": "RATE_LIMITED"}), 429


STATS_SECRET = os.environ.get("STATS_SECRET")


@app.route("/api/ask-parse", methods=["POST"])
@limiter.limit(UPLOAD_RATE_LIMIT)
def ask_parse():
    """
    Takes a natural-language question plus the dataset's schema (not the
    actual data) and returns a structured query spec. The frontend then
    EXECUTES that spec deterministically against the data it already has
    client-side -- this endpoint never sees or computes the real answer.
    """
    body = request.get_json(silent=True) or {}
    question = (body.get("question") or "").strip()
    column_types = body.get("column_types") or {}
    labels = body.get("labels") or {}

    if not question:
        return jsonify({"error": "No question provided"}), 400
    if not column_types:
        return jsonify({"error": "No column information provided"}), 400

    spec = parse_question(question, column_types, labels)
    return jsonify(spec)


@app.route("/api/stats", methods=["GET"])
def stats():
    # Simple secret-key protection -- this is usage data about your app,
    # not something random visitors to your URL should be able to read.
    # Set STATS_SECRET as an env var and pass it as ?key=... to view this.
    if STATS_SECRET:
        if request.args.get("key") != STATS_SECRET:
            return jsonify({"error": "Unauthorized"}), 401
    return jsonify(get_stats())


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=os.environ.get("FLASK_DEBUG", "false") == "true")
